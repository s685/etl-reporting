# ********************************************************************************
# PURPOSE: it is a python script to generate an single Excel file from a SQL query with header or footer logic.
# and script is designed to handle grouped data
# claims decied appeals, claims pending appeals
# ********************************************************************************
# Create an Excel and CSV file with the following format:
# 1. Each column is a separate field in the Excel and CSV file
#    a. If a column is not a string, it's converted to a string
#    b. If a column is null, it's converted to an empty string
#    c. If a column is NaN, it's converted to an empty string
#    d. If the format is still not correct, it is recommended to convert in the SQL query
# 2. Columns are written to the file in the order specified in the config file
#    a. Each row is written to a new line in the file
#    b. Each line is terminated with a newline character
# ********************************************************************************
# NOTES AND RECOMMENDATIONS:
# 1. There's no header or footer logic in this script.
#    Headers are more applicable in a CSV file. The reason is some columns are just 1 character wide
#    and it's not clear what the column represents. In a CSV file, the header can be added to the file
#    Use a UNION ALL to append the footer in the query. If more is needed, a new .py script can be created
# 2. The file is sorted by the SQL query, not by the script
#    This is easily managed in the report definition config file
# 3. There's a delimiter between columns in the CSV file
#    The delimiter is a comma by default, but can be changed in the config file
# ********************************************************************************
# Dependencies:
# 1. Python 3.6 or higher
# 2. snowflake-snowpark-python (via SnowparkConnector)
# 3. datamart_analytics (connector, models, environment)

#************************************
# Standard library imports
import argparse
import logging
import os
import sys
import time
import yaml  # type: ignore[import-untyped]
import pandas as pd
from datamart_analytics.connector.snowpark_connector import SnowparkConnector
from datamart_analytics.models.custom_models import DatamartTable
from datamart_analytics.tools.datamart_utils import create_target_credentials
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from openpyxl.styles import PatternFill, Border, Side, Color
from datetime import datetime

logging.basicConfig(
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%s',
    level=logging.INFO
)

VALID_EXTENSIONS = ['.xlsx']


class FileWriter:
    """Class to write data to an Excel or CSV file.

    parameters:
    output_path: str
        The path where the output file will be saved
    output_file: str
        The name of the output file
    """

    def __init__(self, params):
        self.output_path = params["output_path"]
        self.output_file = params["output_file"]
        self.max_column_width = params["max_column_width"]
        self.sheet_header_font = params["sheet_header_font"]
        self.table_header_font = params["table_header_font"]
        self.table_data_font = params["table_data_font"]
        self.border_to_row = params["border_to_row"]
        self.carrier_name = params["carrier_name"]
        self.report_name = params["report_name"]
        self.report_start_dt = params["report_start_dt"]
        self.report_end_dt = params["report_end_dt"]
        self.report_run_dt = params["report_run_dt"]
        self.report_as_of_run_dt = params["report_as_of_run_dt"]
        self.header = params["header"]
        self.footer = params["footer"]
        self.border_to_row = params["border_to_row"]
        self.dollar_columns = params["dollar_columns"]
        self.specific_column_widths = params["specific_column_widths"]
        self.positive_dollar_format = "${:,.2f}"
        self.negative_dollar_format = "(${:,.2f})"

# write data to an excel file
    def write_to_excel(self, data, ws, current_page, total_pages):
        """ Write data to an Excel file.
        parameters:
            data: pandas DataFrame
            The data to be written to the file
        """

        table_headers = data.columns
        last_column = data.shape[1]
        current_row = 1

        # add report header
        if self.header:
            self.add_header(ws, current_row, last_column, current_page, total_pages)
            current_row += 5

        name, size, bold, color, alignment, wrap_text, fill_color, fill_type = self.set_cell_properties(self.table_header_font)

        # adding table headers
        for col, header in enumerate(table_headers, start=1):
            cell = ws.cell(row=current_row, column=col)
            cell.value = header
            cell.font = Font(name=name, size=size, bold=bold, color=color)
            cell.alignment = Alignment(horizontal=alignment, wrap_text=wrap_text)
            cell.fill = PatternFill(fill_type=fill_type, fgColor=fill_color)

        # set column widths
        logging.info("Setting column widths")
        self.set_column_widths(ws, self.max_column_width)
        self.set_specific_column_widths(data, ws)

        data_rows = data.values.tolist()

        for row in data_rows:
            ws.append(row)
            current_row += 1

    def apply_dollar_format(self, data, ws):
        # Apply dollar format to the columns in the data frame if they are in the dollar_columns list
        if isinstance(data, pd.DataFrame):
            for column in self.dollar_columns:
                logging.info(f"Applying dollar format to column: {column}")
                if column in data.columns:
                    col_idx = data.columns.get_loc(column) + 1  # Get the column index (1-based)
                    column_letter = get_column_letter(col_idx)
                    for cell in ws[column_letter]:
                        cell.number_format = '$#,##0.00'
        else:
            data = self.positive_dollar_format.format(data) if data > 0 else self.negative_dollar_format.format(abs(data))
        return data

    def apply_sorting(self, grouped_data, sorting_columns):
        if sorting_columns is not None:
            grouped_data = grouped_data.sort_values(by=sorting_columns, ascending=True)
        return grouped_data

    def apply_border(self, ws, current_row, last_column, border_to_row):
        if border_to_row['border_to_table_headers']:
            self.apply_border_to_row(ws, current_row, last_column, border_to_row)
        return current_row + 1

    def apply_border_to_row(self, ws, current_row, last_column, border_to_row):
        start_color = border_to_row['start_color']
        end_color = border_to_row['end_color']
        fill_type = border_to_row['fill_type']
        thin_border = Border(top=Side(style='thin'), bottom=Side(style='thin'))

        #current_row += 1
        ws.row_dimensions[current_row].height = 1
        for col_num in range(1, last_column + 1):
            cell = ws.cell(row=current_row, column=col_num)
            cell.border = thin_border
            cell.fill = PatternFill(start_color=Color(rgb=start_color), end_color=Color(rgb=end_color), fill_type=fill_type)

        current_row += 1
        return current_row

    def add_header(self, ws, current_row, last_column, current_page, total_pages):
        # add report header
        name, size, bold, color, alignment, wrap_text, fill_color, fill_type = self.set_cell_properties(self.sheet_header_font)

        for row in range(current_row, current_row + 3):

            cell = ws.cell(row=row, column=1)
            if row == current_row:
                cell.value = self.carrier_name
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_column // 2)

                # Add "timestamp" on the same row as carrier_name
                time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                time_info = f"Executed On: {time}"

                cell_offset = last_column // 2
                time_cell = ws.cell(row=row, column=cell_offset + 1)
                time_cell.value = time_info
                ws.merge_cells(start_row=row, start_column=cell_offset + 1, end_row=row, end_column=last_column)
                time_cell.font = Font(name=name, size=size, bold=bold, color=color)
                time_cell.alignment = Alignment(horizontal='right', wrap_text=wrap_text)
                time_cell.fill = PatternFill(fill_type=fill_type, fgColor=fill_color)

            elif row == current_row + 1:
                cell.value = self.report_name
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_column // 2)

                # Add "Page 1 of 1" on the same row as report_name
                page_info = f"Page {current_page} of {total_pages}"
                # page += 1
                cell_offset = last_column // 2
                page_cell = ws.cell(row=row, column=cell_offset + 1)
                page_cell.value = page_info
                ws.merge_cells(start_row=row, start_column=cell_offset + 1, end_row=row, end_column=last_column)
                page_cell.font = Font(name=name, size=size, bold=bold, color=color)
                page_cell.alignment = Alignment(horizontal='right', wrap_text=wrap_text)
                page_cell.fill = PatternFill(fill_type=fill_type, fgColor=fill_color)
            else:
                if self.report_start_dt and self.report_end_dt:
                    start_date = datetime.strptime(self.report_start_dt, '%Y-%m-%d %H:%M:%S.%f').strftime("%m/%d/%Y")
                    end_date = datetime.strptime(self.report_end_dt, '%Y-%m-%d %H:%M:%S.%f').strftime("%m/%d/%Y")
                    cell.value = f"For Dates: {start_date} To {end_date}"
                else:
                    report_date = datetime.strptime(self.report_run_dt, '%Y-%m-%d %H:%M:%S.%f').strftime("%m/%d/%Y")
                    cell.value = f"Report as Date: {report_date}"
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_column)
                cell.font = Font(name=name, size=size, bold=bold, color=color)
                cell.alignment = Alignment(horizontal=alignment, wrap_text=wrap_text)
                cell.fill = PatternFill(fill_type=fill_type, fgColor=fill_color)

        # set the column widths in the excel file
    def set_column_widths(self, ws, max_column_width):
        for col in range(1, ws.max_column + 1):
            col_index = get_column_letter(col)
            ws.column_dimensions[col_index].width = self.max_column_width

    def set_specific_column_widths(self, data, ws):
        # Set column widths based on the YAML configuration
        if isinstance(data, pd.DataFrame):
            if self.specific_column_widths is not None:
                logging.info("Setting specific column widths")
                for column in self.specific_column_widths:
                    logging.info(f"Setting column width for column {column}")
                    clmn = column['column']
                    wdth = column['width']
                    ws.column_dimensions[clmn].width = wdth

    def set_cell_properties(self, font):
        name = font['name']
        size = font['size']
        bold = font['bold']
        color = font['color']
        wrap_text = font['wrap_text']
        fill_color = font['fill_color']
        fill_type = font['fill_type']
        alignment = font['alignment']

        return name, size, bold, color, alignment, wrap_text, fill_color, fill_type


class Datapreprocessor:
    """Class to handle fetching and processing data from a database.

    parameters:
    connector: SnowparkConnector
        The Snowpark connector (must be used within context manager)
    database: str
        The name of the Snowflake database
    schema: str
        The name of the Snowflake schema
    pre_sql_query: str
        The SQL query to set session variables
    """

    def __init__(self, connector: SnowparkConnector, database: str, schema: str, pre_sql_query: str):
        self.connector = connector
        self.database = database
        self.schema = schema
        self.pre_sql_query = pre_sql_query

        logging.info("Using Snowpark connection")
        logging.info("Active Database.Schema is " + self.database + "." + self.schema)

    def fetch_data(self, table, exclude_columns, filter_rows, sorting_columns):
        """Fetch data from the snowflake database using SnowparkConnector."""

        # set session variables using pre_sql_query
        for statement in self.pre_sql_query.split('\n'):
            if statement.strip():  # ensure the statement is not empty
                self.connector.execute_query(statement, lazy=False)
                logging.info(f"Executed statement: {statement}")

        # fetch data from the tables
        columns = ','.join(['*'] if not exclude_columns else [f'* exclude("{col}")' for col in exclude_columns])
        if filter_rows:
            query = f"SELECT {columns} FROM {table} WHERE {filter_rows}"
        else:
            query = f"SELECT {columns} FROM {table}"

        # sorting_columns: sort the data based on the columns
        if sorting_columns:
            order_by_clause = ', '.join([f'"{col}"' if not col.startswith('"') and not col.endswith('"') else col for col in sorting_columns])
            query += f" ORDER BY {order_by_clause}"

        logging.info(f"Query statement {query}")
        result = self.connector.execute_query(query, lazy=False)

        # Convert list[Row] to DataFrame
        if result is None or len(result) == 0:
            df = pd.DataFrame()
        else:
            df = pd.DataFrame([row.as_dict() for row in result])

        logging.info(f"Data fetched from {table}")
        return df


def validate_report_configextension(report):
    """ validate the report configuration yml file extension"""
    base, ext = os.path.splitext(report)
    return f"{base}.yml" if not ext else report


def load_report_config(report):
    """load the report configuration yml file"""
    with open(report, 'r') as file:
        return yaml.safe_load(file)


def validate_report(report):
    """validate the report configuration yml file"""

    # Check if the report configuration file is empty
    if not report:
        logging.error(f"Error: {report} configuration file is empty.")
        sys.exit(1)

    # Check if the report configuration file is a dictionary
    if not isinstance(report, dict):
        logging.error(f"Error: {report} configuration file is not a dictionary.")
        sys.exit(1)

    # Check if the report configuration file has the required keys:
    # carrier_name and report_name, tables_list(for multiple worksheets), sheetnames(for multiple worksheets), header(for multiple worksheets but some reports may not have header)
    # pre_sql_query (for session variables)
    # sorting_columns (for sorting the data)
    for key in ['carrier_name', 'report_name', 'tables', 'pre_sql_query']:
        if key not in report:
            logging.error(f"Error: {key} key is missing in the report configuration file.")
            sys.exit(1)


def parse_and_validate_args():
    """parse and validate command line arguments"""
    # Assuming argparse.ArgumentParser() is a function that returns an ArgumentParser object
    parser = argparse.ArgumentParser("Required arguments; report, database, schema, output_path, output_file")
    parser.add_argument("report", help="The name of the report config file or extract to create, e.g.,mcas.yml", type=str)
    parser.add_argument("database", help="The name of the Snowflake database, e.g., DEV_SNOWFLAKE_WAREHOUSE", type=str)
    parser.add_argument("schema", help="The name of the Snowflake schema, e.g., BUSINESS_VAULT", type=str)
    parser.add_argument("output_path", help="The path where the output file will be saved, e.g., c:/workspace/", type=str)
    parser.add_argument("output_file", help="The name of the output file, e.g., mcas.xlsx", type=str)
    parser.add_argument("carrier_name", help="The name of the carrier, e.g., ALLIANZ_ADMIN_088", type=str)
    parser.add_argument(
        "--warehouse",
        help="Snowflake warehouse name (or set SNOWFLAKE_WAREHOUSE env var)",
        type=str,
        default=os.getenv("SNOWFLAKE_WAREHOUSE"),
    )

    # # add optinoal arguments : as_of_run_dt, report_start_date, report_end_date, report_run_dt
    parser.add_argument("--as_of_run_dt", help="The ASOF month for the extract, e.g., 12/31/2023", type=str)
    parser.add_argument("--report_start_dt", help="The start date for the report, e.g., 01/01/2023", type=str)
    parser.add_argument("--report_end_dt", help="The end date for the report, e.g., 12/31/2023", type=str)
    parser.add_argument("--report_run_dt", help="The run date for the report, e.g., 12/31/2023", type=str)

    args = parser.parse_args()

    # validate the arguments
    if not args.report:
        raise ValueError("Report name is required")
    if not args.database:
        raise ValueError("Database name is required")
    if not args.schema:
        raise ValueError("Schema name is required")
    if not args.output_path:
        raise ValueError("Output path is required")
    if not args.output_file:
        raise ValueError("Output file name is required")
    if not args.carrier_name:
        raise ValueError("Carrier name is required")
    if not args.warehouse:
        logging.error("Error: Warehouse is required. Set --warehouse or SNOWFLAKE_WAREHOUSE env var.")
        sys.exit(1)

    if not os.path.isfile(args.report):
        logging.error(f"Error: {args.report} configuration file is not a valid file.")
        sys.exit(1)

    if not os.path.isdir(args.output_path):
        logging.error(f"Error: {args.output_path} is not a valid path.")
        sys.exit(1)

    ext = os.path.splitext(args.output_file)[1]
    if ext not in VALID_EXTENSIONS:
        logging.warning(f"WARNING: {args.output_file} does not have a standard file extension.")
    return args


def main():
    """main function"""
    # Assuming parse_and_validate_args() is a function that returns command line arguments
    args = parse_and_validate_args()
    print(args.report)
    # validate the report configuration yml file extension (.yml)
    report_validation = validate_report_configextension(args.report)

    # load the report configuration yml file
    report = load_report_config(report_validation)

    # validate the report configuration yml file keys
    validate_report(report)

    # read the configuration file mandatory keys
    carrier_name = report['carrier_name']
    report_name = report['report_name']
    tables_config = report['tables']
    # sheetnames = report['sheetnames']
    pre_sql_query = report['pre_sql_query'].format(carrier_name=args.carrier_name, as_of_run_dt=args.as_of_run_dt, report_start_dt=args.report_start_dt, report_end_dt=args.report_end_dt)

    # optional keys in config file
    date_columns = report.get('date_columns', None)
    dollar_columns = report.get('dollar_columns', None)
    grouping_column = report.get('grouping_column', None)
    column_widths = report.get('column_widths', None)
    sheet_header_font = report.get('sheet_header_font', None)
    table_header_font = report.get('table_header_font', None)
    table_data_font = report.get('table_data_font', None)
    border_to_row = report.get('border_to_row', None)
    max_column_width = report.get('max_column_width', None)
    report_start_date = report.get('report_start_date', None)
    report_end_date = report.get('report_end_date', None)
    as_of_run_dt = report.get('as_of_run_dt', None)
    report_run_dt = report.get('report_run_dt', None)

    header = report.get('header', None)
    footer = report.get('footer', None)
    # sorting_columns = report.get('sorting_columns', None)
    # exclude_columns = report.get('exclude_columns', None)
    # filter_rows = report.get('filter_rows', None)

    # optional arguments from command line
    report_start_dt = args.report_start_dt if args.report_start_dt else None
    report_end_dt = args.report_end_dt if args.report_end_dt else None
    report_as_of_run_dt = args.as_of_run_dt if args.as_of_run_dt else None
    report_run_dt = args.report_run_dt if args.report_run_dt else None

    # create workbook and write data to an excel file
    datamart_table = DatamartTable(
        name="datafeed",
        source_database=args.database,
        source_schema=args.schema,
        target_database=args.database,
        target_schema=args.schema,
        target_warehouse=args.warehouse,
        carrier_name=args.carrier_name,
    )
    credentials = create_target_credentials(datamart_table)

    wb = Workbook()
    del wb['Sheet']
    current_page = 0
    total_pages = len(tables_config)

    with SnowparkConnector(credentials) as connector:
        dp = Datapreprocessor(
            connector=connector,
            database=args.database,
            schema=args.schema,
            pre_sql_query=pre_sql_query,
        )

        for table_config in tables_config:
            table = table_config['table']
            sheet_name = table_config['sheet_name']
            exclude_columns = table_config.get('excluding_columns', None)
            filter_rows = table_config.get('filter_rows', None)
            sorting_columns = table_config.get('sorting_columns', None)
            specific_column_widths = table_config.get('column_widths', None)
            date_columns = report.get('date_columns', None)
            dollar_columns = report.get('dollar_columns', None)
            sub_group_sum_label = report.get('sub_group_sum_label', None)
            column_widths = report.get('column_widths', None)
            sheet_header_font = report.get('sheet_header_font', None)
            table_header_font = report.get('table_header_font', None)
            table_data_font = report.get('table_data_font', None)
            border_to_row = report.get('border_to_row', None)
            max_column_width = report.get('max_column_width', None)

            params = {
                'output_path': args.output_path,
                'output_file': args.output_file,
                'max_column_width': max_column_width,
                'sheet_header_font': sheet_header_font,
                'table_header_font': table_header_font,
                'table_data_font': table_data_font,
                'carrier_name': carrier_name,
                'report_name': report_name,
                'report_start_dt': report_start_dt,
                'report_end_dt': report_end_dt,
                'report_as_of_run_dt': report_as_of_run_dt,
                'report_run_dt': report_run_dt,
                'header': header,
                'footer': footer,
                'sub_group_sum_label': sub_group_sum_label,
                'border_to_row': border_to_row,
                'dollar_columns': dollar_columns,
                'specific_column_widths': specific_column_widths
            }

            # Fetch and validate the data from Snowflake
            df = dp.fetch_data(table, exclude_columns, filter_rows, sorting_columns)

            current_page = current_page + 1
            ws = wb.create_sheet(title=sheet_name)
            writer = FileWriter(params)
            writer.write_to_excel(df, ws, current_page, total_pages)

            # Apply dollar formatting the the worksheet
            if dollar_columns is not None:
                writer.apply_dollar_format(df, ws)

    wb.save(os.path.join(args.output_path, args.output_file))
    logging.info(f"Output will be saved to: {args.output_path}{args.output_file}")


if __name__ == '__main__':
    start_time = time.time()
    main()
    end_time = time.time()

    execution_time = end_time - start_time
    hours = int(execution_time // 3600)
    minutes = int((execution_time % 3600) // 60)
    seconds = int((execution_time % 60))
    milliseconds = int((execution_time % 1) * 1000)

    print(f'***************************************************')
    print(f'Execution Time For Generating the Feed: {hours} hr {minutes} min {seconds} sec {milliseconds} ms'.center(100))
    print(f'***************************************************')

    logging.info(f'Execution Time For Generating the Feed: {hours} hr {minutes} min {seconds} sec {milliseconds} ms')
