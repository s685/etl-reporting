# ******************************************************************************
# PURPOSE:  python script to generate an Excel file for the following reports:
#   1. claims paid activity report
#   2. claims paid by age and service type report
#   3. claims paid report
# These Reports has many exceptions so created a single script to generate all the reports
# ******************************************************************************
#   Create an Excel and CSV file with the following format:
#       1. Each column is a separate field in the Excel and CSV file
#           a. If a column is not a string, it's converted to a string
#           b. If a column is null, it's converted to an empty string
#           c. If a column is NaN, it's converted to an empty string
#           d. If the format is still not correct, it is recommended to convert in the SQL query
#       2. Columns are written to the file in the order specified in the config file
#           a. Each row is written to a new line in the file
#           b. Each line is terminated with a newline character
# ******************************************************************************
# NOTES AND RECOMMENDATIONS:
#   1. There's no header or footer logic in this script.
#           Headers are more applicable in a CSV file.  The reason is some columns are just 1 character wide
#           and it's not clear what the column represents.  In a CSV file, the header can be added to the file
#           Use a UNION ALL to append the footer in the query.  If more is needed, a new .py script can be created
#   2. The file is sorted by the SQL query, not by the script
#           This is easily managed in the report definition config file
#   3. There's a delimiter between columns in the CSV file
#           The delimiter is a comma by default, but can be changed in the config file
#
# ******************************************************************************
# Dependencies:
#   1. Python 3.6 or higher
#   2. snowflake-sqlalchemy
#   3. snowflake-connector-python
#   4. pandas
#   5. numpy
#   6. argparse
#   7. pyarrow-10.0.1  (newer versions of pyarrow are not compatible with snowflake-connector-python)
#   8. openpyxl
# ******************************************************************************
# Run Command:
#   Arguments:
#       args[0]  = Path to the report definition file
#       args[1]  = Snowflake database name
#       args[2]  = Snowflake schema name
#       args[3]  = Path to the folder where the file will be saved
#       args[4]  = Name of the output file
#       args[5]  = Extract ASOF month
#
#   python EXCEL_CSV_WRITER.py args[0] args[1] args[2] args[3] args[4] args[5]
#   Ex. python EXCEL_CSV_WRITER.py RPT_POLICY_HOLDER_DEMOGRAPHICS DEV_SNOWFLAKE_WAREHOUSE BUSINESS_VAULT d:/workspace/ policyholder_demographics.xlsx 12/31/2023
# Success:
#   1. The file is created in the specified folder location
#   2. The script logs the following informative messages:
#       Initializing connection to <snowflake account>
#       query: <query string>
#       Formatting data for Excel and CSV
#       Writing to <filePath + fileName>
#       File written successfully
#   3. The script exits with status 0 (success)
# ******************************************************************************

# Standard library imports
import argparse
import logging
import os
import sys
import time
import yaml  # type: ignore[import-untyped]
import pandas as pd
from datamart_analytics.connector.snowpark_connector import SnowparkConnector
from datamart_analytics.models.custom_models import DatamartTable
from datamart_analytics.tools.datamart_utils import create_target_credentials
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from openpyxl.styles import PatternFill, Border, Side, Color
from datetime import datetime

logging.basicConfig(
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S',
    level=logging.INFO
)


VALID_EXTENSIONS = ['.xlsx']


class FileWriter:
    """Class to write data to an Excel or CSV file.

    parameters:
    output_path: str
        The path where the output file will be saved
    output_file: str
        The name of the output file
    """

    def __init__(self, params):
        self.output_path = params["output_path"]
        self.output_file = params["output_file"]
        self.sheetnames = params["sheetnames"]
        self.max_column_width = params["max_column_width"]
        self.sorting_columns = params["sorting_columns"]
        self.sheet_header_font = params["sheet_header_font"]
        self.table_header_font = params["table_header_font"]
        self.table_data_font = params["table_data_font"]
        self.carrier_name = params["carrier_name"]
        self.report_name = params["report_name"]
        self.report_start_dt = params["report_start_dt"]
        self.report_end_dt = params["report_end_dt"]
        self.report_run_dt = params["report_run_dt"]
        self.report_as_of_run_dt = params["report_as_of_run_dt"]
        self.header = params["header"]
        self.footer = params["footer"]
        self.grouping_column = params["grouping_column"]

    # write data to a excel file
    def write_to_file(self, data):
        """
        Write data to an Excel

        parameters:
        data: pandas DataFrame
            The data to be written to the file
        """
        _, ext = os.path.splitext(self.output_file)
        if ext == '.xlsx':
            self.write_to_excel(data)
        else:
            logging.error(f"Error: {self.output_file} does not have a valid extension.")
            sys.exit(1)

    # write data to an excel file
    def write_to_excel(self, data):
        """
        Write data to an Excel file.

        parameters:
        data: pandas DataFrame
            The data to be written to the file
        """
        wb = Workbook()
        del wb['Sheet']

        ws = wb.create_sheet(title=self.sheetnames)

        table_headers = data.columns
        last_column = data.shape[1]
        current_row = 1

        # add report header
        if self.header:
            self.add_header(ws, current_row, last_column)
            current_row += 5

        name, size, bold, color, alignment, wrap_text, fill_color, fill_type = self.set_cell_properties(self.table_header_font)

        # adding table headers
        for col, header in enumerate(table_headers, start=1):
            cell = ws.cell(row=current_row, column=col)
            cell.value = header
            cell.font = Font(name=name, size=size, bold=bold, color=color)
            cell.alignment = Alignment(horizontal=alignment, wrap_text=wrap_text)
            cell.fill = PatternFill(fill_type=fill_type, fgColor=fill_color)

        # set column widths
        self.set_column_widths(ws, self.max_column_width)

        if data.empty:
            # add no data available message to the excel file if the data is empty (in one row covering the length of the table)
            # align the text to the center of the cell with a font size of 11 and bold
            ws.merge_cells(start_row=current_row + 2, start_column=1, end_row=current_row + 4, end_column=last_column)
            cell = ws.cell(row=current_row + 2, column=1)
            cell.value = 'No data available'
            cell.font = Font(name='Arial', size=11, bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            wb.save(os.path.join(self.output_path, self.output_file))
            return

        # add table data, get the data from the dataframe for each group
        for group in data[self.grouping_column].unique():
            group_data = data[data[self.grouping_column] == group].sort_values(by=self.sorting_columns)

        print(group_data)
        # add group total
        if self.report_name == 'Claims Paid Report':
            group_total = group_data.groupby(by=self.grouping_column).agg({'Claimants':'sum', 'Amount Paid':'sum'}).reset_index()
            group_total['Avg Paid Per Claimant'] = (group_data['Amount Paid'].sum()) / (group_data['Claimants'].sum())
        else:
            group_total = group_data.groupby(by=self.grouping_column).sum().reset_index()
            group_total = group_total.reindex(columns=group_data.columns, fill_value='')
            group_total[self.sorting_columns] = ''

        # add group total to the group data
        group_data = pd.concat([group_data, group_total], ignore_index=False)
        group_data.reset_index(drop=True, inplace=True)

        group_data[self.grouping_column] = group_data.apply(lambda row: group if row.name == group_data.index[0] else '', axis=1)
        # print(group_data)

        total_rows = len(group_data)
        for row_index, row in enumerate(group_data.itertuples(index=False), start=1):
            current_row += 1
            for col, value in enumerate(row, start=1):
                cell = ws.cell(row=current_row, column=col)
                cell.value = value
                cell.font = Font(name='Arial', size=8, bold=False, color='000000')
                cell.alignment = Alignment(horizontal='right', wrap_text=wrap_text)

                if row_index == total_rows:
                    cell.font = Font(name=name, size=size, bold=bold, color=color)
                    cell.alignment = Alignment(horizontal='right', wrap_text=wrap_text)

        # add total row for the entire data
        if self.report_name == 'Claims Paid Activity':
            data_total = data.sum().to_frame().T
            data_total = data_total.reindex(columns=data.columns, fill_value='')
            data_total[self.grouping_column] = 'Totals'
            data_total[self.sorting_columns] = ''
            self.add_total_row(ws, current_row, data_total)

        elif self.report_name == 'Claims Paid by Age and Service Type':
            # group the data by sorting columns and sum across numeric columns to get the total for each unique value in the sorting columns
            data_category_total = data.groupby(by=self.sorting_columns).sum().reset_index()
            data_category_total = data_category_total.reindex(columns=data.columns, fill_value='').sort_values(by=self.sorting_columns)
            data_category_total[self.grouping_column] = 'Totals for All Groups'
            data_category_total[self.grouping_column] = group_data.apply(lambda row: 'Totals for All Groups' if row.name == group_data.index[0] else '', axis=1)

            current_row += 1
            current_row = self.add_category_total_row(ws, current_row, data_category_total)
            current_row += 1
            data_total = data.agg({'AMOUNT PAID <65':'sum', 'AMOUNT PAID 65-74':'sum', 'AMOUNT PAID 75-79':'sum', 'AMOUNT PAID 80-84':'sum'})
            data_total = data_total.reindex(columns=data.columns, fill_value='')
            data_total[self.grouping_column] = ''
            data_total[self.sorting_columns] = ''
            self.add_total_row(ws, current_row, data_total)

        elif self.report_name == 'Claims Paid Report':
            # group the data by sorting columns and sum across numeric columns to get the total for each unique value in the sorting columns
            data_category_total = data.groupby(by=self.sorting_columns).agg({'Claimants':'sum', 'Amount Paid':'sum', 'Avg Paid Per Claimant':'sum'})
            print(data_category_total)

            data_category_total = data_category_total.reindex(columns=data.columns, fill_value='').sort_values(by=self.sorting_columns)
            data_category_total[self.grouping_column] = 'Totals for All Groups'
            data_category_total[self.grouping_column] = group_data.apply(lambda row: 'Totals for All Groups' if row.name == group_data.index[0] else '', axis=1)

            current_row += 1
            current_row = self.add_category_total_row(ws, current_row, data_category_total)
            current_row += 1
            data_total = data.agg({'Claimants':'sum', 'Amount Paid':'sum'}).to_frame().T
            data_total['Avg Paid Per Claimant'] = (data['Amount Paid'].sum()) / (data['Claimants'].sum())
            data_total = data_total.reindex(columns=data.columns, fill_value='')
            data_total[self.grouping_column] = ''
            data_total[self.sorting_columns] = ''
            self.add_total_row(ws, current_row, data_total)

        wb.save(os.path.join(self.output_path, self.output_file))

    def add_category_total_row(self, ws, current_row, data_total):
        # add total row to the excel file and format the total row
        name, size, bold, color, alignment, wrap_text, fill_color, fill_type = self.set_cell_properties(self.table_header_font)
        for row_index, row in enumerate(data_total.itertuples(index=False), start=1):
            current_row += 1
            for col, value in enumerate(row, start=1):
                cell = ws.cell(row=current_row, column=col)
                cell.value = value
                cell.font = Font(name=name, size=size, bold=bold, color=color)
                cell.alignment = Alignment(horizontal='right', wrap_text=wrap_text)

        return current_row

    def add_total_row(self, ws, current_row, data_total):
        # add total row to the excel file and format the total row
        name, size, bold, color, alignment, wrap_text, fill_color, fill_type = self.set_cell_properties(self.table_header_font)
        for row_index, row in enumerate(data_total.itertuples(index=False), start=1):
            current_row += 1
            for col, value in enumerate(row, start=1):
                cell = ws.cell(row=current_row, column=col)
                cell.value = value
                cell.font = Font(name=name, size=size, bold=bold, color=color)
                cell.alignment = Alignment(horizontal='right', wrap_text=wrap_text)

    def add_header(self, ws, current_row, last_column):
        # add report header
        name, size, bold, color, alignment, wrap_text, fill_color, fill_type = self.set_cell_properties(self.sheet_header_font)

        # Validate fill_type for openpyxl
        valid_fill_types = {
            'none', 'solid', 'darkGray', 'mediumGray', 'lightGray', 'gray0625', 'gray125',
            'lightDown', 'lightUp', 'lightHorizontal', 'lightVertical', 'darkDown', 'darkUp',
            'darkHorizontal', 'darkVertical', 'lightTrellis', 'darkTrellis', 'lightGrid', 'darkGrid'
        }

        if fill_type not in valid_fill_types:
            fill_type = 'solid'

        total_pages = len(self.sheetnames)
        for row in range(current_row, current_row + 3):
            cell = ws.cell(row=row, column=1)
            if row == current_row:
                cell.value = self.carrier_name
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_column // 2)

                # Add "timestamp" on the same row as carrier_name
                time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                time_info = f"Executed On:{time}"

                cell_offset = last_column // 2
                time_cell = ws.cell(row=row, column=cell_offset + 1)
                time_cell.value = time_info
                ws.merge_cells(start_row=row, start_column=cell_offset + 1, end_row=row, end_column=last_column)
                time_cell.font = Font(name=name, size=size, bold=bold, color=color)
                time_cell.alignment = Alignment(horizontal='right', wrap_text=wrap_text)
                time_cell.fill = PatternFill(fill_type=fill_type, fgColor=fill_color)

            elif row == current_row + 1:
                cell.value = self.report_name
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_column // 2)

                # Add "Page 1 of 1" on the same row as report_name
                page_info = f"Page 1 of 1"
                # page += 1
                cell_offset = last_column // 2
                page_cell = ws.cell(row=row, column=cell_offset + 1)
                page_cell.value = page_info
                ws.merge_cells(start_row=row, start_column=cell_offset + 1, end_row=row, end_column=last_column)
                page_cell.font = Font(name=name, size=size, bold=bold, color=color)
                page_cell.alignment = Alignment(horizontal='right', wrap_text=wrap_text)
                page_cell.fill = PatternFill(fill_type=fill_type, fgColor=fill_color)

            else:
                if self.report_start_dt and self.report_end_dt:
                    start_date = datetime.strptime(self.report_start_dt, '%Y-%m-%d %H:%M:%S.%f').strftime("%m/%d/%Y")
                    end_date = datetime.strptime(self.report_end_dt, '%Y-%m-%d %H:%M:%S.%f').strftime("%m/%d/%Y")
                    cell.value = f"For Dates: {start_date} To {end_date}"
                else:
                    report_date = datetime.strptime(self.report_as_of_run_dt, '%Y-%m-%d %H:%M:%S.%f').strftime("%m/%d/%Y")
                    cell.value = f"Report as Date: {report_date}"

                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_column)

            cell.font = Font(name=name, size=size, bold=bold, color=color)
            cell.alignment = Alignment(horizontal=alignment, wrap_text=wrap_text)
            cell.fill = PatternFill(fill_type=fill_type, fgColor=fill_color)

    # set the column widths in the excel file
    def set_column_widths(self, ws, max_column_width):
        for col in range(1, ws.max_column + 1):
            col_index = get_column_letter(col)
            ws.column_dimensions[col_index].width = self.max_column_width

    def set_cell_properties(self, font):
        name = font['name']
        size = font['size']
        bold = font['bold']
        color = font['color']
        wrap_text = font['wrap_text']
        fill_color = font['fill_color']
        fill_type = font['fill_type']
        alignment = font['alignment']

        return name, size, bold, color, alignment, wrap_text, fill_color, fill_type


class Datapreprocessor:
    """Class to handle fetching and processing data from a database.

    parameters:
    connector: SnowparkConnector
        The Snowpark connector (must be used within context manager)
    database: str
        The name of the Snowflake database
    schema: str
        The name of the Snowflake schema
    pre_sql_query: str
        The SQL query to set session variables
    """

    def __init__(self, connector: SnowparkConnector, database: str, schema: str, pre_sql_query: str):
        self.connector = connector
        self.database = database
        self.schema = schema
        self.pre_sql_query = pre_sql_query

        logging.info("Using Snowpark connection")
        logging.info("Active Database.Schema is " + self.database + "." + self.schema)

    def fetch_data(self, table: str, exclude_columns, filter_rows, sorting_columns):
        """Fetch data from the snowflake database using SnowparkConnector."""
        # set session variables using pre_sql_query
        for statement in self.pre_sql_query.split('\n'):
            if statement.strip():  # ensure the statement is not empty
                self.connector.execute_query(statement, lazy=False)
                logging.info(f"Executed statement: {statement}")

        # fetch data from the tables
        columns = ','.join(['*'] if not exclude_columns else [f'* exclude("{col}")' for col in exclude_columns])
        if filter_rows:
            query = f"SELECT {columns} FROM {table} WHERE {filter_rows}"
        else:
            query = f"SELECT {columns} FROM {table}"

        # sorting_columns: sort the data based on the columns
        if sorting_columns:
            order_by_clause = ', '.join([f'"{col}"' if not col.startswith('"') and not col.endswith('"') else col for col in sorting_columns])
            query += f" ORDER BY {order_by_clause}"

        logging.info(f"Query statement {query}")
        result = self.connector.execute_query(query, lazy=False)

        # Convert list[Row] to DataFrame
        if result is None or len(result) == 0:
            df = pd.DataFrame()
        else:
            df = pd.DataFrame([row.as_dict() for row in result])

        logging.info(f"Data fetched from {table}")
        return df


def validate_report_configextension(report):
    """validate the report configuration yml file extension"""
    base, ext = os.path.splitext(report)
    return f"{base}.yml" if not ext else report


def load_report_config(report):
    """load the report configuration yml file"""
    with open(report, 'r') as file:
        return yaml.safe_load(file)


def validate_report(report):
    """validate the report configuration yml file"""

    # Check if the report configuration file is empty
    if not report:
        logging.error(f"Error: {report} configuration file is empty.")
        sys.exit(1)

    # Check if the report configuration file is a dictionary
    if not isinstance(report, dict):
        logging.error(f"Error: {report} configuration file is not a dictionary.")
        sys.exit(1)

    # Check if the report configuration file has the required keys :
    # carrier_name and report_name,tables_list(for multiple worksheets), sheetnames(for multiple worksheets), header(for multiple)
    # pre_sql_query (for session variables)
    # sorting_columns (for sorting the data)
    for key in ['carrier_name', 'report_name', 'tables_list', 'sheetnames', 'pre_sql_query']:
        if key not in report:
            logging.error(f"Error: {key} key is missing in the report configuration file.")
            sys.exit(1)


def parse_and_validate_args():
    """parse and validate command line arguments"""
    # Assuming argparse.ArgumentParser() is a function that returns an ArgumentParser object
    parser = argparse.ArgumentParser("Required arguments: report, database, schema, output_path, output_file")
    parser.add_argument("report", help="The name of the report config file or extract to create, e.g., .mcas.yml", type=str)
    parser.add_argument("database", help="The name of the Snowflake database, e.g., DEV_SNOWFLAKE_WAREHOUSE", type=str)
    parser.add_argument("schema", help="The name of the Snowflake schema, e.g., BUSINESS_VAULT", type=str)
    parser.add_argument("output_path", help="The path where the output file will be saved, e.g., c:/workspace/", type=str)
    parser.add_argument("output_file", help="The name of the output file, e.g., mcas.xlsx", type=str)
    parser.add_argument("carrier_name", help="The name of the carrier, e.g., ALLIANZ_ADMIN_008", type=str)
    parser.add_argument(
        "--warehouse",
        help="Snowflake warehouse name (or set SNOWFLAKE_WAREHOUSE env var)",
        type=str,
        default=os.getenv("SNOWFLAKE_WAREHOUSE"),
    )

    # # add optional arguments : as_of_run_dt, report_start_date, report_end_date, report_run_dt
    parser.add_argument("--as_of_run_dt", help="The ASOF month for the extract, e.g., 12/31/2023", type=str)
    parser.add_argument("--report_start_dt", help="The start date for the report, e.g., 01/01/2023", type=str)
    parser.add_argument("--report_end_dt", help="The end date for the report, e.g., 12/31/2023", type=str)
    parser.add_argument("--report_run_dt", help="The run date for the report, e.g., 12/31/2023", type=str)

    args = parser.parse_args()

    # validate the arguments
    if not args.report:
        raise ValueError("Report name is required")
    if not args.database:
        raise ValueError("Database name is required")
    if not args.schema:
        raise ValueError("Schema name is required")
    if not args.output_path:
        raise ValueError("Output path is required")
    if not args.output_file:
        raise ValueError("Output file name is required")
    if not args.carrier_name:
        raise ValueError("Carrier name is required")
    if not args.as_of_run_dt:
        raise ValueError("as_of_run_dt is required")
    if not args.warehouse:
        logging.error("Error: Warehouse is required. Set --warehouse or SNOWFLAKE_WAREHOUSE env var.")
        sys.exit(1)

    if not os.path.isfile(args.report):
        logging.error(f"Error: {args.report} configuration file is not a valid file.")
        sys.exit(1)

    if not os.path.isdir(args.output_path):
        logging.error(f"Error: {args.output_path} is not a valid path.")
        sys.exit(1)

    ext = os.path.splitext(args.output_file)[1]
    if ext not in VALID_EXTENSIONS:
        logging.warning(f"WARNING: {args.output_file} does not have a standard file extension.")

    return args


def main():
    """main function"""
    # Assuming parse_and_validate_args() is a function that returns command line arguments
    args = parse_and_validate_args()

    print(args.report)
    # validate the report configuration yml file extension (.yml)
    report_validation = validate_report_configextension(args.report)

    # load the report configuration yml file
    report = load_report_config(report_validation)

    # validate the report configuration yml file keys
    validate_report(report)

    # read the configuration file mandatory keys
    carrier_name = report['carrier_name']
    report_name = report['report_name']
    tables_list = report['tables_list']
    sheetnames = report['sheetnames']
    pre_sql_query = report['pre_sql_query'].format(carrier_name=args.carrier_name, as_of_run_dt=args.as_of_run_dt, report_start_dt=args.report_start_dt, report_end_dt=args.report_end_dt, report_run_dt=args.report_run_dt)

    # optional keys in config file
    # ---
    # date_columns, dollar_columns, grouping_column, column_widths, sheet_header_font,
    # table_header_font, table_data_font, border_to_row, max_column_width, report_start_date, report_end_date, as_of_run_dt, report_run_dt
    # ---
    header = report.get('header', None)
    footer = report.get('footer', None)
    sorting_columns = report.get('sorting_columns', None)
    exclude_columns = report.get('exclude_columns', None)
    filter_rows = report.get('filter_rows', None)
    date_columns = report.get('date_columns', None)
    dollar_columns = report.get('dollar_columns', None)
    grouping_column = report.get('grouping_column', None)
    grp_sum = report.get('grp_sum', None)
    grp_cnt_label = report.get('grp_cnt_label', 'None')
    grp_sum_label = report.get('grp_sum_label', None)
    tot_cnt_label = report.get('tot_cnt_label', 'None')
    tot_sum_label = report.get('tot_sum_label', None)
    column_widths = report.get('column_widths', None)
    sheet_header_font = report.get('sheet_header_font', None)
    table_header_font = report.get('table_header_font', None)
    table_data_font = report.get('table_data_font', None)
    border_to_row = report.get('border_to_row', None)
    max_column_width = report.get('max_column_width', None)

    # for claim paid reversals
    additional_tables = report.get('additional_tables', None)
    additional_sheetnames = report.get('additional_sheetnames', None)

    # optional arguments from command line
    report_start_dt = args.report_start_dt if args.report_start_dt else None
    report_end_dt = args.report_end_dt if args.report_end_dt else None
    report_as_of_run_dt = args.as_of_run_dt if args.as_of_run_dt else None
    report_run_dt = args.report_run_dt if args.report_run_dt else None

    # Create credentials and Snowpark connector
    datamart_table = DatamartTable(
        name="datafeed",
        source_database=args.database,
        source_schema=args.schema,
        target_database=args.database,
        target_schema=args.schema,
        target_warehouse=args.warehouse,
        carrier_name=args.carrier_name,
    )
    credentials = create_target_credentials(datamart_table)

    params = {
        'output_path': args.output_path,
        'output_file': args.output_file,
        'sheetnames': sheetnames,
        'max_column_width': max_column_width,
        'sorting_columns': sorting_columns,
        'sheet_header_font': sheet_header_font,
        'table_header_font': table_header_font,
        'table_data_font': table_data_font,
        'carrier_name': carrier_name,
        'report_name': report_name,
        'report_start_dt': report_start_dt,
        'report_end_dt': report_end_dt,
        'report_as_of_run_dt': report_as_of_run_dt,
        'report_run_dt': report_run_dt,
        'header': header,
        'footer': footer,
        'grouping_column': grouping_column,
        'grp_sum': grp_sum,
        'grp_cnt_label': grp_cnt_label,
        'grp_sum_label': grp_sum_label,
        'tot_cnt_label': tot_cnt_label,
        'tot_sum_label': tot_sum_label,
        'border_to_row': border_to_row
    }

    with SnowparkConnector(credentials) as connector:
        dp = Datapreprocessor(
            connector=connector,
            database=args.database,
            schema=args.schema,
            pre_sql_query=pre_sql_query,
        )

        # Fetch and validate the data from Snowflake
        df = dp.fetch_data(tables_list, exclude_columns, filter_rows, sorting_columns)

        # Create an instance of the FileWriter class and write data to file
        writer = FileWriter(params)
        status = writer.write_to_file(df)

        logging.info(f"Output will be saved to: {args.output_path}{args.output_file}")

        # Only for Claims Paid Report, add additional worksheet(s) if configured
        if report_name == 'Claims Paid Report' and additional_tables and additional_sheetnames:
            # Support both string and list for additional_tables/sheetnames:
            if isinstance(additional_tables, str):
                additional_tables = [additional_tables]
            if isinstance(additional_sheetnames, str):
                additional_sheetnames = [additional_sheetnames]
            from openpyxl import load_workbook
            wb_path = os.path.join(args.output_path, args.output_file)
            wb = load_workbook(wb_path)
            for table, sheet in zip(additional_tables, additional_sheetnames):
                df_add = dp.fetch_data(table, exclude_columns, filter_rows, None)  # No sorting for additional tables
                ws = wb.create_sheet(title=sheet)
                # Add header using the same logic as add_header
                writer.add_header(ws, 1, df_add.shape[1])
                current_row = 6  # after header (3 rows + 2 blank + 1 for table header)
                # Add table headers
                name, size, bold, color, alignment, wrap_text, fill_color, fill_type = writer.set_cell_properties(table_header_font)
                for col, header in enumerate(df_add.columns, start=1):
                    cell = ws.cell(row=current_row, column=col)
                    cell.value = header
                    cell.font = Font(name=name, size=size, bold=bold, color=color)
                    cell.alignment = Alignment(horizontal=alignment, wrap_text=wrap_text)
                    cell.fill = PatternFill(fill_type=fill_type, fgColor=fill_color)
                writer.set_column_widths(ws, max_column_width)
                # Add data rows
                for row in df_add.itertuples(index=False):
                    current_row += 1
                    for col, value in enumerate(row, start=1):
                        cell = ws.cell(row=current_row, column=col)
                        cell.value = value
                        cell.font = Font(name='Arial', size=8, bold=False, color='000000')
                        cell.alignment = Alignment(horizontal='right', wrap_text=wrap_text)

                # Add summation row for 'Amount Reversed' column only, with label 'Total' before it
                if not df_add.empty and 'Amount Reversed' in df_add.columns:
                    amount_col_idx = list(df_add.columns).index('Amount Reversed') + 1  # 1-based index for openpyxl
                    current_row += 1
                    # Fill blanks up to the label cell
                    for col_idx in range(1, amount_col_idx):
                        cell = ws.cell(row=current_row, column=col_idx)
                        if col_idx == amount_col_idx - 1:
                            cell.value = 'Total'
                            cell.font = Font(name=name, size=size, bold=True, color=color)
                            cell.alignment = Alignment(horizontal='right', wrap_text=wrap_text)
                        else:
                            cell.value = ''
                    # Write the sum in the correct column
                    cell = ws.cell(row=current_row, column=amount_col_idx)
                    cell.value = df_add['Amount Reversed'].sum()
                    cell.font = Font(name=name, size=size, bold=True, color=color)
                    cell.alignment = Alignment(horizontal='right', wrap_text=wrap_text)
                    # Fill blanks for any columns after
                    for col_idx in range(amount_col_idx + 1, df_add.shape[1] + 1):
                        cell = ws.cell(row=current_row, column=col_idx)
                        cell.value = ''
            wb.save(wb_path)
            logging.info(f"Additional worksheet(s) with totals added to: {wb_path}")


if __name__ == '__main__':

    start_time = time.time()
    main()
    end_time = time.time()

    execution_time = end_time - start_time
    hours = int(execution_time // 3600)
    minutes = int((execution_time % 3600) // 60)
    seconds = int(execution_time % 60)
    milliseconds = int((execution_time % 1) * 1000)

    print(f'{"="*100}')
    print(f'Execution Time For Generating the Feed: {hours} hr {minutes} min {seconds} sec {milliseconds} ms'.center(100))
    print(f'{"="*100}')

    logging.info(f'Execution Time For Generating the Feed: {hours} hr {minutes} min {seconds} sec {milliseconds} ms')
