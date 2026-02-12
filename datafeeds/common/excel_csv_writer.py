# ******************************************************************************
# PURPOSE:
#   Python script to generate an Excel or CSV file from a SQL query.
#   Generates files without any header or footer in the script's direct output.
#
# OUTPUT FILE FORMAT (Excel and CSV):
#   1. Each column from the SQL query result is a separate field.
#      a. If a column's data type is not a string, it is converted to a string.
#      b. NULL values are converted to an empty string.
#      c. NaN values are converted to an empty string.
#      d. If formatting issues persist, handle conversion in the SQL query.
#   2. Columns are written in the order specified in the configuration file.
#      a. Each row is written to a new line.
#      b. Each line is terminated with a newline character.
#
# NOTES AND RECOMMENDATIONS:
#   1. Headers/Footers: The script does not implement header or footer logic.
#      Headers are more applicable in CSV files. Some columns may be 1 character
#      wide. Use UNION ALL in the SQL query to append a footer if needed.
#      For CSV, the header can be added via the report definition config.
#   2. Sorting: Content is sorted by the SQL query, not by this script
#      (managed in the report definition configuration file).
#   3. Delimiter: CSV delimiter defaults to comma; can be customized in config.
#
# DEPENDENCIES:
#   Python 3.6+, snowflake-snowpark-python (via SnowparkConnector),
#   datamart_analytics (connector, models, tools), pandas, numpy, argparse,
#   pyarrow=10.0.1, openpyxl
#
# RUN COMMAND:
#   python EXCEL_CSV_WRITER.py args[0] args[1] args[2] args[3] args[4] args[5]
#   Ex. python EXCEL_CSV_WRITER.py <report_config> <database> <schema> <output_path> <output_file> <carrier_name>
#
# SUCCESS:
#   1. File is created in the specified folder location.
#   2. Script logs: connection init, query, formatting, writing path, success.
#   3. Script exits with status 0.
# ******************************************************************************

import argparse
import logging
import os
import sys
import time
import yaml  # type: ignore[import-untyped]
import datetime

import pandas as pd
from datamart_analytics.connector.snowpark_connector import SnowparkConnector
from datamart_analytics.models.custom_models import DatamartTable
from datamart_analytics.tools.datamart_utils import create_target_credentials

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

logging.basicConfig(
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S',
    level=logging.INFO
)

VALID_EXTENSIONS = ['.xlsx', '.csv']

REQUIRED_REPORT_KEYS = [
    'carrier_name', 'report_name', 'tables_list', 'sheetnames',
    'reporting_headers', 'pre_sql_query'
]


def validate_report_configextension(report):
    """Validate the report configuration yml file extension."""
    base, ext = os.path.splitext(report)
    return f"{base}.yml" if not ext else report


def load_report_config(report):
    """Load the report configuration yml file."""
    with open(report, 'r') as file:
        return yaml.safe_load(file)


def validate_report(report):
    """Validate the report configuration yml file."""
    if not report:
        logging.error(f"Error: {report} configuration file is empty.")
        sys.exit(1)
    if not isinstance(report, dict):
        logging.error(f"Error: {report} configuration file is not a dictionary.")
        sys.exit(1)
    for key in REQUIRED_REPORT_KEYS:
        if key not in report:
            logging.error(f"Error: {key} key is missing in the report configuration file.")
            sys.exit(1)


def apply_date_format(df, date_columns, date_format='%Y-%m-%d'):
    """Apply date formatting to the date columns in the DataFrame."""
    for column in date_columns:
        if column in df.columns:
            df[column] = df[column].dt.strftime(date_format)
    return df


def parse_and_validate_args():
    """Parse and validate command line arguments."""
    parser = argparse.ArgumentParser(
        description="Required arguments: report, database, schema, output_path, output_file, carrier_name"
    )
    parser.add_argument('report', help="Report config file or extract name, e.g. mcas.yml", type=str)
    parser.add_argument('database', help="Snowflake database, e.g. DEV_SNOWFLAKE_WAREHOUSE", type=str)
    parser.add_argument('schema', help="Snowflake schema, e.g. BUSINESS_VAULT", type=str)
    parser.add_argument('output_path', help="Path where output file will be saved, e.g. c:/workspace/", type=str)
    parser.add_argument('output_file', help="Output file name, e.g. mcas.xlsx", type=str)
    parser.add_argument('carrier_name', help="Carrier name, e.g. ALLIANZ_ADMIN_08B", type=str)
    parser.add_argument(
        '--warehouse',
        help="Snowflake warehouse name (or set SNOWFLAKE_WAREHOUSE env var)",
        type=str,
        default=os.environ.get('SNOWFLAKE_WAREHOUSE'),
    )
    parser.add_argument('--as_of_run_dt', help="ASOF month for extract, e.g. 12/31/2023", type=str, default=None)
    parser.add_argument('--report_start_dt', help="Start date for report, e.g. 01/01/2023", type=str, default=None)
    parser.add_argument('--report_end_dt', help="End date for report, e.g. 12/31/2023", type=str, default=None)
    parser.add_argument('--report_run_dt', help="Run date for report, e.g. 12/31/2023", type=str, default=None)

    args = parser.parse_args()

    if not args.report:
        raise ValueError("Report name is required")
    if not args.database:
        raise ValueError("Database is required")
    if not args.schema:
        raise ValueError("Schema is required")
    if not args.output_path:
        raise ValueError("Output path is required")
    if not args.output_file:
        raise ValueError("Output file is required")
    if not args.carrier_name:
        raise ValueError("Carrier name is required")

    if not os.path.isfile(args.report):
        logging.error(f"Error: {args.report} configuration file is not a valid file.")
        sys.exit(1)
    if not os.path.isdir(args.output_path):
        logging.error(f"Error: {args.output_path} is not a valid path.")
        sys.exit(1)
    if not args.warehouse:
        logging.error("Error: Warehouse is required. Set --warehouse or SNOWFLAKE_WAREHOUSE env var.")
        sys.exit(1)

    _, ext = os.path.splitext(args.output_file)
    if ext.lower() not in VALID_EXTENSIONS:
        logging.warning(f"WARNING: {args.output_file} does not have a standard file extension.")

    return args


def batch_control(connector, log_params):
    """Insert batch control log into Snowflake BATCH_CONTROL table using Snowpark."""
    database = log_params['database']
    schema = log_params['schema']
    try:
        query = f"""INSERT INTO {database}.{schema}.BATCH_CONTROL (
            SEQ_BATCH_CONTROL_ID,
            BATCH_CONTROL_DATE,
            CARRIER_NAME,
            JOB_NAME,
            JOB_START_TIME,
            JOB_END_TIME,
            BATCH_STATUS,
            STATUS_REASON,
            OUTPUT_FILE_NAME,
            RECORD_COUNT
        ) VALUES (
            SEQ_BATCH_CONTROL_ID.NEXTVAL,
            CURRENT_TIMESTAMP(),
            '{log_params['carrier_name']}',
            '{log_params['report_name']}',
            '{log_params['start_time']}',
            '{log_params['end_time']}',
            '{log_params['status']}',
            '{log_params['status_reason']}',
            '{log_params['output_file']}',
            {log_params['record_count']}
        )"""
        logging.info("Inserting batch control log")
        connector.execute_query(query, lazy=False)
        logging.info("Batch control log inserted successfully")
    except Exception as e:
        logging.error(f"Error: {e}")
        sys.exit(1)


class FileWriter:
    """Class to write data to an Excel or CSV file.

    parameters:
        output_path: str - The path where the output file will be saved.
        output_file: str - The name of the output file.
    """

    def __init__(self, output_path, output_file, sheetnames, max_column_width,
                 sorting_column, table_header_font, table_data_font, carriage_return):
        self.output_path = output_path
        self.output_file = output_file
        self.sheetnames = sheetnames
        self.max_column_width = max_column_width
        self.sorting_column = sorting_column
        self.table_header_font = table_header_font
        self.table_data_font = table_data_font
        self.carriage_return = carriage_return

    def write_to_file(self, data):
        """Write data to an Excel or CSV file.

        parameters:
            data: pandas DataFrame - The data to be written to the file.
        """
        _, ext = os.path.splitext(self.output_file)
        if ext == '.xlsx':
            return self.write_to_excel(data)
        elif ext == '.csv':
            return self.write_to_csv(data)
        else:
            logging.error(f"Error: {self.output_file} does not have a valid extension.")
            sys.exit(1)

    def write_to_excel(self, data):
        """Write data to an Excel file.

        parameters:
            data: pandas DataFrame - The data to be written to the file.
        """
        wb = Workbook()
        del wb['Sheet']
        ws = wb.create_sheet(title=self.sheetnames[0] if isinstance(self.sheetnames, list) else self.sheetnames)

        table_headers = data.columns
        current_row = 1

        name, size, bold, color, alignment, wrap_text, fill_color, fill_type = self.set_cell_properties(
            self.table_header_font
        )
        # adding table headers
        for col, header in enumerate(table_headers, start=1):
            cell = ws.cell(row=current_row, column=col)
            cell.value = header
            cell.font = Font(name=name, size=size, bold=bold, color=color)
            cell.alignment = Alignment(horizontal=alignment, wrap_text=wrap_text)
            cell.fill = PatternFill(fill_type=fill_type, fgColor=fill_color)

        # set column widths
        self.set_column_widths(ws, self.max_column_width)

        data_rows = data.values.tolist()
        for row in data_rows:
            ws.append(row)

        self.wb = wb
        wb.save(os.path.join(self.output_path, self.output_file))
        return 'SUCCESS'

    def write_to_csv(self, data):
        """Write data to a CSV file.

        parameters:
            data: pandas DataFrame - The data to be written to the file.
        """
        path = os.path.join(self.output_path, self.output_file)
        if self.carriage_return:
            data.to_csv(path, index=False, header=True, lineterminator='\r\n')
        else:
            data.to_csv(path, index=False, header=True)
        return 'SUCCESS'

    def set_column_widths(self, ws, max_column_width):
        """Set the column widths in the Excel file."""
        for col in range(1, ws.max_column + 1):
            col_index = get_column_letter(col)
            ws.column_dimensions[col_index].width = max_column_width or self.max_column_width

    def set_cell_properties(self, font):
        """Extract font/cell properties from a font configuration dict."""
        if not font or not isinstance(font, dict):
            return 'Calibri', 11, False, '000000', 'general', False, '00000000', 'none'
        name = font.get('name', 'Calibri')
        size = font.get('size', 11)
        bold = font.get('bold', False)
        color = font.get('color', '000000')
        wrap_text = font.get('wrap_text', False)
        fill_color = font.get('fill_color', '00000000')
        fill_type = font.get('fill_type', 'none')
        alignment = font.get('alignment', 'general')
        return name, size, bold, color, alignment, wrap_text, fill_color, fill_type


class Datapreprocessor:
    """Fetch and process data from Snowflake using SnowparkConnector.

    parameters:
        connector: SnowparkConnector
            The Snowpark connector (must be used within context manager).
        database: str - The name of the Snowflake database.
        schema: str - The name of the Snowflake schema.
        pre_sql_query: str - SQL query to set session variables.
        tables_list: list - List of tables to fetch (one for single sheet, multiple for multiple worksheets).
    """

    def __init__(self, connector, database, schema, pre_sql_query, tables_list):
        self.connector = connector
        self.database = database
        self.schema = schema
        self.pre_sql_query = pre_sql_query or ''
        self.tables_list = tables_list if isinstance(tables_list, list) else [tables_list]
        self.table = self.tables_list[0] if self.tables_list else None

        logging.info("Using Snowpark connection")
        logging.info("Active Database.Schema is " + self.database + "." + self.schema)

    def fetch_data(self, exclude_columns=None, filter_rows=None, sorting_columns=None):
        """Fetch data from the Snowflake database using SnowparkConnector."""
        exclude_columns = exclude_columns or []

        for statement in self.pre_sql_query.split('\n'):
            if statement.strip():
                self.connector.execute_query(statement, lazy=False)
                logging.info(f"Executed statement: {statement}")

        columns = ','.join(
            ['*'] if not exclude_columns else [f'* EXCLUDE("{col}")' for col in exclude_columns]
        )
        if filter_rows:
            query = f"SELECT {columns} FROM {self.table} WHERE {filter_rows}"
        else:
            query = f"SELECT {columns} FROM {self.table}"

        if sorting_columns:
            order_by_clause = ', '.join(
                f'"{col}"' if not (col.startswith('"') and col.endswith('"')) else col
                for col in sorting_columns
            )
            query += f" ORDER BY {order_by_clause}"

        logging.info(f"Query statement {query}")
        result = self.connector.execute_query(query, lazy=False)

        if result is None or len(result) == 0:
            df = pd.DataFrame()
        else:
            df = pd.DataFrame([row.as_dict() for row in result])
        logging.info(f"Data fetched from {self.table}")
        return df


def main():
    """Main function."""
    start_time = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    args = None
    report_name = None
    df = None
    status = 'FAILED'
    status_reason = 'Unknown error'
    end_time = start_time

    try:
        args = parse_and_validate_args()
        report_validation = validate_report_configextension(args.report)
        report = load_report_config(report_validation)
        validate_report(report)

        carrier_name = report['carrier_name']
        report_name = report['report_name']
        tables_list = report['tables_list']
        sheetnames = report['sheetnames']
        reporting_headers = report['reporting_headers']
        pre_sql_query = report['pre_sql_query'].format(
            carrier_name=args.carrier_name,
            as_of_run_dt=args.as_of_run_dt or '',
            report_start_dt=args.report_start_dt or '',
            report_end_dt=args.report_end_dt or '',
            report_run_dt=args.report_run_dt or ''
        )

        # optional keys in config file
        sorting_columns = report.get('sorting_columns', None)
        exclude_columns = report.get('exclude_columns', None)
        filter_rows = report.get('filter_rows', None)
        date_columns = report.get('date_columns', None)
        dollar_columns = report.get('dollar_columns', None)
        grouping_column = report.get('grouping_column', None)
        column_widths = report.get('column_widths', None)
        sheet_header_font = report.get('sheet_header_font', None)
        table_header_font = report.get('table_header_font', None)
        table_data_font = report.get('table_data_font', None)
        border_to_row = report.get('border_to_row', None)
        max_column_width = report.get('max_column_width', None)
        carriage_return = report.get('carriage_return', None)

        datamart_table = DatamartTable(
            name="datafeed",
            source_database=args.database,
            source_schema=args.schema,
            target_database=args.database,
            target_schema=args.schema,
            target_warehouse=args.warehouse,
            carrier_name=args.carrier_name,
        )
        credentials = create_target_credentials(datamart_table)

        with SnowparkConnector(credentials) as connector:
            try:
                dp = Datapreprocessor(
                    connector=connector,
                    database=args.database,
                    schema=args.schema,
                    pre_sql_query=pre_sql_query,
                    tables_list=tables_list,
                )
                df = dp.fetch_data(
                    exclude_columns=exclude_columns,
                    filter_rows=filter_rows,
                    sorting_columns=sorting_columns,
                )
                writer = FileWriter(
                    args.output_path,
                    args.output_file,
                    sheetnames,
                    max_column_width,
                    sorting_columns,
                    table_header_font,
                    table_data_font,
                    carriage_return
                )
                status = writer.write_to_file(df)
                logging.info(f"Writing to {args.output_path}{args.output_file}")
                end_time = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                status_reason = 'File written successfully'
            except Exception as e:
                status_reason = str(e)
                logging.error(f"Error: {e}")
                end_time = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                status = 'FAILED'

            log_params = {
                'carrier_name': args.carrier_name,
                'report_name': report_name,
                'start_time': start_time,
                'end_time': end_time,
                'status': status,
                'status_reason': status_reason,
                'output_file': args.output_file,
                'record_count': len(df) if df is not None else 0,
                'database': args.database,
                'schema': args.schema,
            }
            print(log_params)
            batch_control(connector, log_params)

    except Exception as e:
        status_reason = str(e)
        logging.error(f"Error: {e}")
        end_time = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        status = 'FAILED'


if __name__ == '__main__':
    start_time = time.time()
    main()
    end_time = time.time()
    execution_time = end_time - start_time
    hours = int(execution_time // 3600)
    minutes = int((execution_time % 3600) // 60)
    seconds = int(execution_time % 60)
    milliseconds = int((execution_time % 1) * 1000)
    print('****************************************************************************************************'.center(100))
    print(f'Execution Time For Generating the Feed: {hours} hr {minutes} min {seconds} sec {milliseconds} ms'.center(100))
    print('****************************************************************************************************'.center(100))
    logging.info(
        f'Execution Time For Generating the Feed: {hours} hr {minutes} min {seconds} sec {milliseconds} ms'
    )
