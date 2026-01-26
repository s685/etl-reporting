"""
PDF Table Extractor
Parses tables from PDF files and converts them to Excel or CSV format.

This script supports multiple PDF table extraction libraries for maximum compatibility:
- pdfplumber (default, best for most PDFs)
- tabula-py (Java-based, good for complex layouts)
- camelot-py (best for lattice/stream tables)

Usage:
    python pdf_table_extractor.py --input report.pdf --output report.xlsx
    python pdf_table_extractor.py --input report.pdf --output report.csv --format csv
    python pdf_table_extractor.py --input report.pdf --output report.xlsx --library tabula
"""

import argparse
import logging
import sys
from pathlib import Path
from typing import List, Literal
import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Configure logging
logging.basicConfig(
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S',
    level=logging.INFO
)
logger = logging.getLogger(__name__)


class PDFTableExtractor:
    """Extract tables from PDF files using multiple extraction methods."""
    
    SUPPORTED_LIBRARIES = ['pdfplumber', 'tabula', 'camelot']
    SUPPORTED_FORMATS = ['excel', 'csv']
    
    def __init__(
        self, 
        input_pdf: str, 
        output_file: str, 
        library: str = 'pdfplumber',
        output_format: str = 'excel',
        combine_tables: bool = True,
        detail_only: bool = True,
        min_detail_rows: int = 10
    ):
        """
        Initialize PDF table extractor.
        
        Args:
            input_pdf: Path to input PDF file
            output_file: Path to output file (Excel or CSV)
            library: Extraction library to use ('pdfplumber', 'tabula', 'camelot')
            output_format: Output format ('excel' or 'csv')
            combine_tables: If True, combine all tables into one (default: True)
            detail_only: If True, extract only detail tables (not summaries) (default: True)
            min_detail_rows: Minimum rows to be considered detail data (default: 10)
        """
        self.input_pdf = Path(input_pdf)
        self.output_file = Path(output_file)
        self.library = library.lower()
        self.output_format = output_format.lower()
        self.combine_tables = combine_tables
        self.detail_only = detail_only
        self.min_detail_rows = min_detail_rows
        
        # Validate inputs
        self._validate_inputs()
    
    def _validate_inputs(self) -> None:
        """Validate input parameters."""
        # Check if PDF exists
        if not self.input_pdf.exists():
            raise FileNotFoundError(f"PDF file not found: {self.input_pdf}")
        
        if not self.input_pdf.suffix.lower() == '.pdf':
            raise ValueError(f"Input file must be a PDF: {self.input_pdf}")
        
        # Check library support
        if self.library not in self.SUPPORTED_LIBRARIES:
            raise ValueError(
                f"Library '{self.library}' not supported. "
                f"Choose from: {', '.join(self.SUPPORTED_LIBRARIES)}"
            )
        
        # Check format support
        if self.output_format not in self.SUPPORTED_FORMATS:
            raise ValueError(
                f"Format '{self.output_format}' not supported. "
                f"Choose from: {', '.join(self.SUPPORTED_FORMATS)}"
            )
        
        # Create output directory if it doesn't exist
        self.output_file.parent.mkdir(parents=True, exist_ok=True)
    
    def extract_tables(self) -> List[pd.DataFrame]:
        """
        Extract all tables from PDF using specified library.
        
        Returns:
            List of DataFrames, one per table found
        
        Raises:
            ImportError: If required library is not installed
            Exception: If extraction fails
        """
        logger.info(f"Extracting tables from: {self.input_pdf}")
        logger.info(f"Using library: {self.library}")
        
        if self.library == 'pdfplumber':
            return self._extract_with_pdfplumber()
        elif self.library == 'tabula':
            return self._extract_with_tabula()
        elif self.library == 'camelot':
            return self._extract_with_camelot()
        else:
            raise ValueError(f"Unsupported library: {self.library}")
    
    def _extract_with_pdfplumber(self) -> List[pd.DataFrame]:
        """Extract tables using pdfplumber library."""
        try:
            import pdfplumber
        except ImportError:
            raise ImportError(
                "pdfplumber not installed. Install with: pip install pdfplumber"
            )
        
        tables = []
        last_headers = None  # Track headers for multi-page tables
        last_num_cols = None
        
        with pdfplumber.open(self.input_pdf) as pdf:
            logger.info(f"PDF has {len(pdf.pages)} pages")
            
            for page_num, page in enumerate(pdf.pages, start=1):
                logger.info(f"Processing page {page_num}/{len(pdf.pages)}")
                
                # Extract tables from page
                page_tables = page.extract_tables()
                
                if page_tables:
                    for table_num, table in enumerate(page_tables, start=1):
                        if table and len(table) > 0:
                            # Check if first row looks like headers or data
                            first_row = table[0]
                            num_cols = len(first_row)
                            
                            # Heuristic: If first row has similar column count and looks like continuation
                            is_continuation = False
                            if last_headers is not None and num_cols == last_num_cols:
                                # Check if first row looks like data (not headers)
                                # Data typically has numbers, dates, or varied content
                                is_continuation = True
                                logger.debug(f"  Detected continuation table on page {page_num}")
                            
                            if is_continuation and last_headers is not None:
                                # Use previous headers, all rows are data
                                df = pd.DataFrame(table, columns=last_headers)
                                logger.debug(f"  Using headers from previous page")
                            else:
                                # First row is headers
                                df = pd.DataFrame(table[1:], columns=table[0])
                                last_headers = table[0]  # Save headers for next page
                                last_num_cols = num_cols
                            
                            # Clean up DataFrame
                            df = self._clean_dataframe(df)
                            
                            # Reset index to avoid duplicate index issues
                            df = df.reset_index(drop=True)
                            
                            # Validate it's a proper table
                            if not self._is_valid_table(df, f"Page {page_num}"):
                                continue
                            
                            # Filter for detail tables only if requested
                            # Pass is_continuation flag to skip row count check for continuation pages
                            if self.detail_only and not self._is_detail_table(df, f"Page {page_num}", is_continuation=is_continuation):
                                continue
                            
                            # Add metadata
                            df.attrs['page'] = page_num
                            df.attrs['table_num'] = table_num
                            
                            tables.append(df)
                            table_type = "DETAIL" if not self.detail_only else ""
                            logger.info(
                                f"  Found {table_type} table {table_num} on page {page_num}: "
                                f"{len(df)} rows x {len(df.columns)} columns"
                            )
        
        if not tables:
            logger.warning("No tables found in PDF")
        
        return tables
    
    def _extract_with_tabula(self) -> List[pd.DataFrame]:
        """Extract tables using tabula-py library."""
        try:
            import tabula
        except ImportError:
            raise ImportError(
                "tabula-py not installed. Install with: pip install tabula-py\n"
                "Note: Also requires Java to be installed"
            )
        
        logger.info("Extracting tables with tabula-py (this may take a moment)...")
        
        # Extract all tables from all pages
        tables = tabula.read_pdf(
            str(self.input_pdf),
            pages='all',
            multiple_tables=True,
            lattice=True,  # Use lattice mode for tables with borders
            stream=False   # Use stream mode for tables without borders if lattice fails
        )
        
        if not tables:
            logger.warning("No tables found. Trying stream mode...")
            tables = tabula.read_pdf(
                str(self.input_pdf),
                pages='all',
                multiple_tables=True,
                lattice=False,
                stream=True
            )
        
        # Clean tables
        cleaned_tables = []
        for idx, df in enumerate(tables, start=1):
            if not df.empty:
                df = self._clean_dataframe(df)
                df.attrs['table_num'] = idx
                cleaned_tables.append(df)
                logger.info(f"Found table {idx}: {len(df)} rows x {len(df.columns)} columns")
        
        return cleaned_tables
    
    def _extract_with_camelot(self) -> List[pd.DataFrame]:
        """Extract tables using camelot-py library."""
        try:
            import camelot
        except ImportError:
            raise ImportError(
                "camelot-py not installed. Install with: pip install camelot-py[cv]\n"
                "Note: Also requires Ghostscript to be installed"
            )
        
        logger.info("Extracting tables with camelot-py...")
        
        # Try lattice mode first (for tables with borders)
        tables = camelot.read_pdf(
            str(self.input_pdf),
            pages='all',
            flavor='lattice'
        )
        
        if len(tables) == 0:
            logger.warning("No tables found with lattice mode. Trying stream mode...")
            tables = camelot.read_pdf(
                str(self.input_pdf),
                pages='all',
                flavor='stream'
            )
        
        # Convert to DataFrames
        dataframes = []
        for idx, table in enumerate(tables, start=1):
            df = table.df
            
            # Use first row as header if it looks like headers
            if len(df) > 0:
                df.columns = df.iloc[0]
                df = df[1:]
                df = df.reset_index(drop=True)
            
            df = self._clean_dataframe(df)
            df.attrs['page'] = table.page
            df.attrs['table_num'] = idx
            df.attrs['accuracy'] = table.accuracy
            
            dataframes.append(df)
            logger.info(
                f"Found table {idx} on page {table.page}: "
                f"{len(df)} rows x {len(df.columns)} columns "
                f"(accuracy: {table.accuracy:.2f}%)"
            )
        
        return dataframes
    
    def _is_detail_table(self, df: pd.DataFrame, debug_info: str = "", is_continuation: bool = False) -> bool:
        """
        Determine if table is detail data (vs summary data).
        
        Detail tables have:
        - More rows (granular/transactional data)
        - More columns (detailed information)
        - Individual records (not aggregated summaries)
        
        Summary tables have:
        - Fewer rows (aggregated data)
        - Often contain "Total", "Summary", "Subtotal" in values
        - Typically smaller
        
        Args:
            df: DataFrame to check
            debug_info: Optional debug info for logging
            is_continuation: If True, skip row count check (continuation of multi-page table)
        
        Returns:
            True if detail table, False if summary table
        """
        # Check 1: Row count (detail tables have more rows)
        # Skip this check for continuation pages (last page might have fewer rows)
        if not is_continuation and len(df) < self.min_detail_rows:
            logger.debug(f"  {debug_info} Identified as SUMMARY (only {len(df)} rows, need {self.min_detail_rows}+)")
            return False
        
        # Check 2: Look for summary keywords in data
        summary_keywords = ['total', 'summary', 'subtotal', 'grand total', 'sum', 'aggregate']
        
        # Convert all data to string and check for keywords
        df_str = df.astype(str).values.flatten()
        df_lower = [str(val).lower() for val in df_str]
        
        # Count how many cells contain summary keywords
        keyword_matches = sum(1 for val in df_lower if any(keyword in val for keyword in summary_keywords))
        
        # If more than 20% of cells contain summary keywords, likely a summary table
        total_cells = len(df_str)
        if total_cells > 0 and (keyword_matches / total_cells) > 0.2:
            logger.debug(f"  {debug_info} Identified as SUMMARY (contains summary keywords)")
            return False
        
        # Check 3: Column count ratio (detail tables usually have more columns)
        # If table has very few columns AND few rows, likely summary
        if len(df.columns) <= 3 and len(df) < self.min_detail_rows * 2:
            logger.debug(f"  {debug_info} Identified as SUMMARY (only {len(df.columns)} columns and {len(df)} rows)")
            return False
        
        logger.debug(f"  {debug_info} Identified as DETAIL ({len(df)} rows x {len(df.columns)} columns)")
        return True
    
    def _is_valid_table(self, df: pd.DataFrame, debug_info: str = "") -> bool:
        """
        Check if DataFrame is a valid table (has headers and data).
        
        Args:
            df: DataFrame to validate
            debug_info: Optional debug info for logging (e.g., "page 1")
        
        Returns:
            True if valid table, False otherwise
        """
        # Must have at least 1 data row
        if len(df) < 1:
            logger.debug(f"  {debug_info} Skipped: No data rows")
            return False
        
        # Must have at least 1 column (relaxed from 2)
        if len(df.columns) < 1:
            logger.debug(f"  {debug_info} Skipped: No columns")
            return False
        
        # Check if it has meaningful column names (not all empty)
        non_empty_cols = [str(col).strip() for col in df.columns if str(col).strip()]
        if len(non_empty_cols) == 0:
            logger.debug(f"  {debug_info} Skipped: All column names empty")
            return False
        
        # Check if it has at least some data
        has_data = not df.empty and df.notna().any().any()
        if not has_data:
            logger.debug(f"  {debug_info} Skipped: No data (all NaN)")
            return False
        
        # Check if table is too small (likely a page number or header)
        total_cells = len(df) * len(df.columns)
        if total_cells < 3:  # At least 3 cells
            logger.debug(f"  {debug_info} Skipped: Too small ({total_cells} cells)")
            return False
        
        return True
    
    def _clean_dataframe(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Clean extracted DataFrame.
        
        Args:
            df: Raw DataFrame from extraction
        
        Returns:
            Cleaned DataFrame
        """
        # Remove completely empty rows
        df = df.dropna(how='all')
        
        # Remove completely empty columns
        df = df.dropna(axis=1, how='all')
        
        # Strip whitespace from string columns
        for col in df.columns:
            try:
                # Try to strip whitespace if column contains string-like data
                df[col] = df[col].astype(str).str.strip()
            except (AttributeError, TypeError):
                # If column can't be converted to string, skip it
                pass
        
        # Replace None/NaN with empty string for string columns
        df = df.fillna('')
        
        # Reset index
        df = df.reset_index(drop=True)
        
        return df
    
    def _combine_tables(self, tables: List[pd.DataFrame]) -> pd.DataFrame:
        """
        Combine multiple tables into one DataFrame.
        
        Handles tables with different columns by using union of all columns.
        
        Args:
            tables: List of DataFrames to combine
        
        Returns:
            Combined DataFrame
        
        Raises:
            ValueError: If no valid tables to combine
        """
        if not tables:
            raise ValueError("No tables to combine - all tables were filtered out")
        
        if len(tables) == 1:
            logger.info(f"Single table: {len(tables[0])} rows x {len(tables[0].columns)} columns")
            return tables[0]
        
        logger.info(f"Combining {len(tables)} tables into one...")
        
        # Filter out any empty DataFrames
        non_empty_tables = [df for df in tables if not df.empty and len(df.columns) > 0]
        
        if not non_empty_tables:
            raise ValueError("All tables are empty after cleaning")
        
        # Get all unique columns across all tables
        all_columns = []
        for df in non_empty_tables:
            all_columns.extend(df.columns.tolist())
        unique_columns = list(dict.fromkeys(all_columns))  # Preserve order
        
        # Reindex all tables to have same columns
        aligned_tables = []
        for idx, df in enumerate(non_empty_tables, 1):
            try:
                # Create a copy to avoid modifying original
                df_copy = df.copy()
                
                # Reset index to ensure uniqueness
                df_copy = df_copy.reset_index(drop=True)
                
                # Add missing columns with empty values
                for col in unique_columns:
                    if col not in df_copy.columns:
                        df_copy[col] = ''
                
                # Reorder columns to match
                df_aligned = df_copy[unique_columns]
                
                # Ensure clean index for concatenation
                df_aligned = df_aligned.reset_index(drop=True)
                
                aligned_tables.append(df_aligned)
                
                logger.debug(f"  Aligned table {idx}: {len(df_aligned)} rows x {len(df_aligned.columns)} columns")
                
            except Exception as e:
                logger.error(f"  Error aligning table {idx}: {e}")
                # Try to continue with other tables
                continue
        
        if not aligned_tables:
            raise ValueError("Failed to align any tables for concatenation")
        
        # Concatenate all tables
        try:
            combined = pd.concat(aligned_tables, ignore_index=True, sort=False)
        except Exception as e:
            logger.error(f"Error during concatenation: {e}")
            logger.error("Attempting alternative concatenation method...")
            
            # Alternative: Convert all to string first
            try:
                string_tables = []
                for df in aligned_tables:
                    df_str = df.astype(str)
                    string_tables.append(df_str)
                combined = pd.concat(string_tables, ignore_index=True, sort=False)
            except Exception as e2:
                raise ValueError(f"Cannot concatenate tables: {e}. Alternative method also failed: {e2}")
        
        logger.info(f"Combined table: {len(combined)} rows x {len(combined.columns)} columns")
        
        return combined
    
    def _format_excel_sheet(self, worksheet, df: pd.DataFrame) -> None:
        """
        Apply professional formatting to Excel worksheet.
        
        Args:
            worksheet: openpyxl worksheet object
            df: DataFrame that was written to the sheet
        """
        # Define styles
        header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        data_font = Font(name='Calibri', size=10)
        data_alignment = Alignment(horizontal='left', vertical='center')
        
        border_style = Border(
            left=Side(style='thin', color='D3D3D3'),
            right=Side(style='thin', color='D3D3D3'),
            top=Side(style='thin', color='D3D3D3'),
            bottom=Side(style='thin', color='D3D3D3')
        )
        
        # Format header row (row 1)
        for col_num, column in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border_style
        
        # Format data rows
        for row_num in range(2, len(df) + 2):
            for col_num in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = border_style
        
        # Auto-fit column widths
        for col_num, column in enumerate(df.columns, 1):
            column_letter = get_column_letter(col_num)
            
            # Calculate max length for column
            max_length = len(str(column))  # Header length
            for value in df[column].astype(str):
                max_length = max(max_length, len(str(value)))
            
            # Set column width (with limits)
            adjusted_width = min(max_length + 2, 50)  # Max width of 50
            worksheet.column_dimensions[column_letter].width = max(adjusted_width, 10)  # Min width of 10
        
        # Freeze header row
        worksheet.freeze_panes = 'A2'
        
        logger.debug(f"  Applied formatting: {len(df)} rows x {len(df.columns)} columns")
    
    def save_to_excel(self, tables: List[pd.DataFrame]) -> None:
        """
        Save tables to Excel file.
        
        If combine_tables=True, all tables go into one sheet.
        If combine_tables=False, each table gets its own sheet.
        
        Args:
            tables: List of DataFrames to save
        
        Raises:
            ValueError: If no valid tables to save
        """
        if not tables:
            raise ValueError("No tables to save - all tables may have been filtered out")
        
        logger.info(f"Saving {len(tables)} table(s) to Excel: {self.output_file}")
        
        with pd.ExcelWriter(self.output_file, engine='openpyxl') as writer:
            if self.combine_tables:
                # Combine all tables into one
                combined_df = self._combine_tables(tables)
                
                # Validate combined table is not empty
                if combined_df.empty or len(combined_df.columns) == 0:
                    raise ValueError("Combined table is empty - no valid data found")
                
                # Write data
                combined_df.to_excel(writer, sheet_name='Combined_Data', index=False)
                
                # Apply formatting
                worksheet = writer.sheets['Combined_Data']
                self._format_excel_sheet(worksheet, combined_df)
                
                logger.info(f"  Saved combined table: {len(combined_df)} rows x {len(combined_df.columns)} columns")
            else:
                # Save each table to separate sheet
                sheets_saved = 0
                for idx, df in enumerate(tables, start=1):
                    # Skip empty DataFrames
                    if df.empty or len(df.columns) == 0:
                        logger.warning(f"  Skipped empty table {idx}")
                        continue
                    
                    # Create sheet name
                    if 'page' in df.attrs:
                        sheet_name = f"Page{df.attrs['page']}_Table{df.attrs.get('table_num', idx)}"
                    else:
                        sheet_name = f"Table_{idx}"
                    
                    # Excel sheet names must be <= 31 characters
                    sheet_name = sheet_name[:31]
                    
                    # Write to Excel
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                    
                    # Apply formatting
                    worksheet = writer.sheets[sheet_name]
                    self._format_excel_sheet(worksheet, df)
                    
                    logger.info(f"  Saved sheet: {sheet_name}")
                    sheets_saved += 1
                
                if sheets_saved == 0:
                    raise ValueError("No valid sheets saved - all tables were empty")
        
        logger.info(f"Successfully saved to: {self.output_file}")
    
    def save_to_csv(self, tables: List[pd.DataFrame]) -> None:
        """
        Save tables to CSV file(s).
        
        If combine_tables=True, all tables saved to one CSV.
        If combine_tables=False and multiple tables, creates separate files.
        
        Args:
            tables: List of DataFrames to save
        """
        if self.combine_tables:
            # Combine all tables into one CSV
            combined_df = self._combine_tables(tables)
            combined_df.to_csv(self.output_file, index=False)
            logger.info(f"Saved combined table to: {self.output_file}")
            logger.info(f"  {len(combined_df)} rows x {len(combined_df.columns)} columns")
        elif len(tables) == 1:
            # Single table - save to specified filename
            tables[0].to_csv(self.output_file, index=False)
            logger.info(f"Saved to: {self.output_file}")
        else:
            # Multiple tables - save with suffixes
            base_name = self.output_file.stem
            parent_dir = self.output_file.parent
            
            for idx, df in enumerate(tables, start=1):
                output_path = parent_dir / f"{base_name}_table_{idx}.csv"
                df.to_csv(output_path, index=False)
                logger.info(f"Saved table {idx} to: {output_path}")
        
        logger.info(f"Successfully saved {len(tables)} table(s)")
    
    def process(self) -> None:
        """Main processing method: extract and save tables."""
        try:
            # Extract tables
            tables = self.extract_tables()
            
            if not tables:
                logger.error("No valid tables found in PDF")
                logger.error("")
                logger.error("Possible reasons:")
                logger.error("  - PDF contains no tables")
                logger.error("  - Tables don't have clear headers/structure")
                if self.detail_only:
                    logger.error(f"  - All tables filtered as SUMMARY (< {self.min_detail_rows} rows)")
                    logger.error("")
                    logger.error("Solutions:")
                    logger.error("  1. Include summary tables: --include-summary")
                    logger.error(f"  2. Lower threshold: --min-detail-rows 5")
                    logger.error("  3. Try different library: --library tabula")
                else:
                    logger.error("  - Try a different extraction library: --library tabula or --library camelot")
                sys.exit(1)
            
            logger.info(f"Total valid tables extracted: {len(tables)}")
            if self.detail_only:
                logger.info(f"(Detail tables only, summaries filtered out)")
            
            # Save to output format
            if self.output_format == 'excel':
                self.save_to_excel(tables)
            elif self.output_format == 'csv':
                self.save_to_csv(tables)
            
            logger.info("Processing complete!")
            
        except ValueError as e:
            logger.error(f"Error: {e}")
            logger.error("")
            if self.detail_only:
                logger.error("Suggestions:")
                logger.error("  1. Try: --include-summary")
                logger.error("  2. Try: --min-detail-rows 5")
            else:
                logger.error("Try using --separate-tables flag or a different extraction library")
            sys.exit(1)
        except Exception as e:
            logger.error(f"Error during processing: {e}", exc_info=True)
            raise


def parse_args() -> argparse.Namespace:
    """Parse command line arguments."""
    parser = argparse.ArgumentParser(
        description='Extract tables from PDF and convert to Excel or CSV',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Extract using pdfplumber (default) to Excel
  python pdf_table_extractor.py --input report.pdf --output report.xlsx
  
  # Extract to CSV format
  python pdf_table_extractor.py --input report.pdf --output report.csv --format csv
  
  # Use tabula library (good for complex layouts)
  python pdf_table_extractor.py --input report.pdf --output report.xlsx --library tabula
  
  # Use camelot library (best for bordered tables)
  python pdf_table_extractor.py --input report.pdf --output report.xlsx --library camelot
        """
    )
    
    parser.add_argument(
        '--input',
        '-i',
        required=True,
        help='Input PDF file path'
    )
    
    parser.add_argument(
        '--output',
        '-o',
        required=True,
        help='Output file path (Excel or CSV)'
    )
    
    parser.add_argument(
        '--library',
        '-l',
        default='pdfplumber',
        choices=['pdfplumber', 'tabula', 'camelot'],
        help='PDF extraction library to use (default: pdfplumber)'
    )
    
    parser.add_argument(
        '--format',
        '-f',
        default='excel',
        choices=['excel', 'csv'],
        help='Output format (default: excel)'
    )
    
    parser.add_argument(
        '--combine-tables',
        action='store_true',
        default=True,
        help='Combine all tables into one (default: True)'
    )
    
    parser.add_argument(
        '--separate-tables',
        action='store_true',
        help='Keep tables separate (one sheet/file per table)'
    )
    
    parser.add_argument(
        '--detail-only',
        action='store_true',
        default=True,
        help='Extract only detail tables, skip summary tables (default: True)'
    )
    
    parser.add_argument(
        '--include-summary',
        action='store_true',
        help='Include both detail and summary tables'
    )
    
    parser.add_argument(
        '--min-detail-rows',
        type=int,
        default=10,
        help='Minimum rows to be considered detail data (default: 10)'
    )
    
    return parser.parse_args()


def main():
    """Main entry point."""
    args = parse_args()
    
    # Auto-detect format from output extension if not specified
    output_ext = Path(args.output).suffix.lower()
    if output_ext in ['.xlsx', '.xls']:
        args.format = 'excel'
    elif output_ext == '.csv':
        args.format = 'csv'
    
    # Determine if tables should be combined
    combine_tables = not args.separate_tables  # Default True unless --separate-tables specified
    
    # Determine if only detail tables should be extracted
    detail_only = not args.include_summary  # Default True unless --include-summary specified
    
    # Create extractor and process
    extractor = PDFTableExtractor(
        input_pdf=args.input,
        output_file=args.output,
        library=args.library,
        output_format=args.format,
        combine_tables=combine_tables,
        detail_only=detail_only,
        min_detail_rows=args.min_detail_rows
    )
    
    extractor.process()


if __name__ == '__main__':
    main()
